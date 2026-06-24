from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from sdgk.core.paths import ensure_students_path


STATUS_FILL = {
    "PASS": "D9EAD3",
    "MATCH": "D9EAD3",
    "BLOCK": "F4CCCC",
    "REVIEW": "FFF2CC",
    "NO_MATCH": "EADCF8",
}


def add_rows(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    for column_index, header in enumerate(headers, start=1):
        width = min(42, max(12, len(str(header)) + 4))
        for cell in ws[get_column_letter(column_index)]:
            width = min(42, max(width, min(40, len(str(cell.value or "")) + 2)))
            if str(cell.value or "") in STATUS_FILL:
                cell.fill = PatternFill("solid", fgColor=STATUS_FILL[str(cell.value)])
        ws.column_dimensions[get_column_letter(column_index)].width = width


def write_plan_workbook(
    *,
    out_path: Path,
    profile: dict[str, Any],
    candidate_pool: dict[str, Any],
    strategy_result: dict[str, Any],
    final_audit: dict[str, Any],
) -> Path:
    safe_path = ensure_students_path(out_path)
    safe_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "说明"
    ws.append(["山东高考志愿方案", "模拟方案" if candidate_pool.get("simulation") else "正式候选"])
    ws.append(["硬闸门", "通过" if final_audit.get("hard_gate_passed") else "未通过"])
    ws.append(["考生", profile.get("name") or profile.get("姓名") or ""])
    ws.append(["年份", profile.get("year") or profile.get("年份") or ""])
    ws.append(["分数", profile.get("score") or profile.get("分数") or ""])
    ws.append(["位次", candidate_pool.get("rank_info", {}).get("rank") or ""])
    ws.append(["说明", "正式提交前必须以山东省教育招生考试院最新官方数据复核。"])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80
    for row in ws.iter_rows():
        row[0].font = Font(bold=True)

    selected = strategy_result.get("selected", [])
    volunteer_rows: list[dict[str, Any]] = []
    for row in selected:
        volunteer_rows.append(
            {
                "序号": row.get("strategy_order"),
                "梯度": row.get("gradient_bucket"),
                "录取概率": row.get("probability"),
                "院校代码": row.get("school_code"),
                "院校名称": row.get("school_name"),
                "专业代码": row.get("major_code"),
                "专业名称": row.get("major_name"),
                "专业标签": row.get("major_family"),
                "省份": row.get("province"),
                "城市": row.get("city"),
                "选科审核状态": row.get("subject_check_status"),
                "地区审核状态": row.get("region_check_status"),
                "program_id": row.get("program_id"),
                "evidence_id": row.get("evidence_id"),
                "source_file": row.get("source_file"),
            }
        )
    add_rows(
        wb.create_sheet("志愿表"),
        ["序号", "梯度", "录取概率", "院校代码", "院校名称", "专业代码", "专业名称", "专业标签", "省份", "城市", "选科审核状态", "地区审核状态", "program_id", "evidence_id", "source_file"],
        volunteer_rows,
    )
    add_rows(
        wb.create_sheet("梯度与风险"),
        ["指标", "值"],
        [
            {"指标": "硬闸门", "值": final_audit.get("hard_gate_passed")},
            {"指标": "保守滑档概率", "值": strategy_result.get("conservative_slip_probability")},
            {"指标": "独立模型滑档概率", "值": strategy_result.get("independent_model_slip_probability")},
            {"指标": "冲稳保垫", "值": str(strategy_result.get("gradient_counts", {}))},
            {"指标": "违规项", "值": "; ".join(strategy_result.get("violations", []))},
        ],
    )
    add_rows(
        wb.create_sheet("专业匹配"),
        ["序号", "院校名称", "专业名称", "专业标签", "匹配偏好", "偏好分"],
        [
            {
                "序号": row.get("strategy_order"),
                "院校名称": row.get("school_name"),
                "专业名称": row.get("major_name"),
                "专业标签": row.get("major_family"),
                "匹配偏好": "、".join(row.get("matched_preferences") or []),
                "偏好分": row.get("preference_fit"),
            }
            for row in selected
        ],
    )
    add_rows(
        wb.create_sheet("地区审核"),
        ["序号", "院校名称", "省份", "城市", "状态", "原因", "证据ID"],
        [
            {
                "序号": row.get("strategy_order"),
                "院校名称": row.get("school_name"),
                "省份": row.get("province"),
                "城市": row.get("city"),
                "状态": row.get("region_check_status"),
                "原因": row.get("region_reason_code"),
                "证据ID": row.get("region_evidence_id"),
            }
            for row in selected
        ],
    )
    add_rows(
        wb.create_sheet("选科审核"),
        ["序号", "院校名称", "专业名称", "状态", "原因", "选科要求", "证据ID", "来源"],
        [
            {
                "序号": row.get("strategy_order"),
                "院校名称": row.get("school_name"),
                "专业名称": row.get("major_name"),
                "状态": row.get("subject_check_status"),
                "原因": row.get("subject_reason_code"),
                "选科要求": row.get("subject_requirement_raw"),
                "证据ID": row.get("subject_evidence_id"),
                "来源": row.get("subject_source_file"),
            }
            for row in selected
        ],
    )
    add_rows(
        wb.create_sheet("证据索引"),
        ["序号", "program_id", "evidence_id", "subject_evidence_id", "region_evidence_id", "source_file"],
        [
            {
                "序号": row.get("strategy_order"),
                "program_id": row.get("program_id"),
                "evidence_id": row.get("evidence_id"),
                "subject_evidence_id": row.get("subject_evidence_id"),
                "region_evidence_id": row.get("region_evidence_id"),
                "source_file": row.get("source_file"),
            }
            for row in selected
        ],
    )
    review_rows = candidate_pool.get("review_rows", []) + strategy_result.get("blocked", [])[:100]
    add_rows(
        wb.create_sheet("人工复核清单"),
        ["院校名称", "专业名称", "选科状态", "地区状态", "原因", "source_file"],
        [
            {
                "院校名称": row.get("school_name"),
                "专业名称": row.get("major_name"),
                "选科状态": row.get("subject_check_status"),
                "地区状态": row.get("region_check_status"),
                "原因": ";".join(row.get("blocked_reasons", [])) or row.get("subject_reason_code") or row.get("region_reason_code"),
                "source_file": row.get("source_file"),
            }
            for row in review_rows[:300]
        ],
    )
    wb.save(safe_path)
    return safe_path
