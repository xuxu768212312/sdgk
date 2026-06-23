#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Audit a volunteer plan against the deterministic subject index."""

from __future__ import annotations

import argparse
import csv
import json
import re
import zipfile
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple
from xml.etree import ElementTree as ET

from subject_eligibility import DEFAULT_DB_PATH, PASS_STATUS, check_eligibility, normalize_subjects


HEADER_ALIASES = {
    "school_code": {"schoolcode", "院校代码", "院校代号", "学校代码", "学校代号"},
    "major_code": {"majorcode", "专业代码", "专业代号"},
    "school_name": {"schoolname", "院校名称", "学校名称", "院校", "学校"},
    "major_name": {"majorname", "专业名称", "专业", "专业类"},
    "order": {"order", "index", "序号", "志愿序号", "志愿号"},
}


def normalize_header(value: Any) -> str:
    return re.sub(r"[\s_：:（）()]+", "", str(value or "").strip().lower())


def find_column(headers: Sequence[str], canonical: str) -> Optional[str]:
    normalized_to_original = {normalize_header(h): h for h in headers}
    for alias in HEADER_ALIASES[canonical]:
        key = normalize_header(alias)
        if key in normalized_to_original:
            return normalized_to_original[key]
    return None


def build_column_map(headers: Sequence[str]) -> Dict[str, Optional[str]]:
    return {key: find_column(headers, key) for key in HEADER_ALIASES}


def read_csv_rows(path: Path) -> List[Dict[str, Any]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        return [dict(row) for row in csv.DictReader(f)]


def read_json_rows(path: Path) -> List[Dict[str, Any]]:
    with path.open(encoding="utf-8") as f:
        payload = json.load(f)
    if isinstance(payload, list):
        return [dict(row) for row in payload]
    if isinstance(payload, dict):
        for key in ("volunteers", "items", "rows", "data"):
            value = payload.get(key)
            if isinstance(value, list):
                return [dict(row) for row in value]
    raise ValueError("JSON input must be a list or contain volunteers/items/rows/data list")


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_xlsx_rows_openpyxl(path: Path) -> List[Dict[str, Any]]:
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [[cell_text(cell.value) for cell in row] for row in ws.iter_rows()]
    return table_rows_to_dicts(rows)


def xlsx_col_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha())
    idx = 0
    for ch in letters:
        idx = idx * 26 + (ord(ch.upper()) - ord("A") + 1)
    return max(0, idx - 1)


def read_xlsx_rows_stdlib(path: Path) -> List[Dict[str, Any]]:
    ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with zipfile.ZipFile(path) as zf:
        names = set(zf.namelist())
        shared: List[str] = []
        if "xl/sharedStrings.xml" in names:
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.findall("a:si", ns):
                parts = [t.text or "" for t in si.findall(".//a:t", ns)]
                shared.append("".join(parts))

        sheet_names = sorted(name for name in names if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"))
        if not sheet_names:
            raise ValueError("XLSX has no worksheet XML")
        sheet_root = ET.fromstring(zf.read(sheet_names[0]))

    rows: List[List[str]] = []
    for row_el in sheet_root.findall(".//a:sheetData/a:row", ns):
        values: List[str] = []
        for c in row_el.findall("a:c", ns):
            ref = c.attrib.get("r", "")
            col_idx = xlsx_col_index(ref)
            while len(values) <= col_idx:
                values.append("")

            cell_type = c.attrib.get("t")
            value = ""
            if cell_type == "s":
                v = c.find("a:v", ns)
                if v is not None and v.text is not None:
                    value = shared[int(v.text)]
            elif cell_type == "inlineStr":
                value = "".join(t.text or "" for t in c.findall(".//a:t", ns))
            else:
                v = c.find("a:v", ns)
                if v is not None and v.text is not None:
                    value = v.text
            values[col_idx] = cell_text(value)
        rows.append(values)
    return table_rows_to_dicts(rows)


def table_rows_to_dicts(rows: List[List[str]]) -> List[Dict[str, Any]]:
    rows = [row for row in rows if any(cell_text(cell) for cell in row)]
    if not rows:
        return []
    headers = [cell_text(cell) for cell in rows[0]]
    result: List[Dict[str, Any]] = []
    for row in rows[1:]:
        item = {}
        for idx, header in enumerate(headers):
            if header:
                item[header] = cell_text(row[idx]) if idx < len(row) else ""
        if any(str(v).strip() for v in item.values()):
            result.append(item)
    return result


def read_xlsx_rows(path: Path) -> List[Dict[str, Any]]:
    try:
        return read_xlsx_rows_openpyxl(path)
    except ImportError:
        return read_xlsx_rows_stdlib(path)


def read_input_rows(path: Path) -> List[Dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(path)
    if suffix == ".json":
        return read_json_rows(path)
    if suffix == ".xlsx":
        return read_xlsx_rows(path)
    raise ValueError("input must be CSV, JSON, or XLSX")


def require_columns(rows: List[Dict[str, Any]]) -> Dict[str, Optional[str]]:
    if not rows:
        raise ValueError("input contains no volunteer rows")
    headers: List[str] = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                headers.append(key)
    column_map = build_column_map(headers)
    required = ["school_code", "major_code", "school_name", "major_name"]
    missing = [name for name in required if not column_map.get(name)]
    if missing:
        raise ValueError("missing required columns: " + ", ".join(missing))
    return column_map


def get_value(row: Dict[str, Any], column_map: Dict[str, Optional[str]], key: str) -> str:
    column = column_map.get(key)
    if not column:
        return ""
    return cell_text(row.get(column))


def audit_rows(
    *,
    rows: List[Dict[str, Any]],
    year: Optional[int],
    edition: Optional[str],
    level: str,
    subjects: str,
    db_path: Path,
) -> Dict[str, Any]:
    column_map = require_columns(rows)
    normalized_subjects, invalid_subjects = normalize_subjects(subjects)
    audited_rows: List[Dict[str, Any]] = []

    for index, row in enumerate(rows, start=1):
        result = check_eligibility(
            year=year,
            edition=edition,
            level=level,
            subjects=subjects,
            school_code=get_value(row, column_map, "school_code"),
            major_code=get_value(row, column_map, "major_code"),
            school_name=get_value(row, column_map, "school_name"),
            major_name=get_value(row, column_map, "major_name"),
            db_path=db_path,
        )
        audited_rows.append(
            {
                "row_number": index,
                "order": get_value(row, column_map, "order") or index,
                "school_code": get_value(row, column_map, "school_code"),
                "school_name": get_value(row, column_map, "school_name"),
                "major_code": get_value(row, column_map, "major_code"),
                "major_name": get_value(row, column_map, "major_name"),
                "subject_check_status": result["status"],
                "eligible": result["eligible"],
                "reason_code": result["reason_code"],
                "message": result["message"],
                "match_type": result["match_type"],
                "evidence_id": result.get("evidence_id"),
                "source_file": result.get("source_file"),
                "evidence": result.get("evidence"),
            }
        )

    counts = Counter(row["subject_check_status"] for row in audited_rows)
    overall_status = PASS_STATUS if counts and counts.get(PASS_STATUS, 0) == len(audited_rows) else "BLOCK_OR_REVIEW"
    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "overall_status": overall_status,
        "hard_gate_passed": overall_status == PASS_STATUS,
        "year": year,
        "edition": edition,
        "level": level,
        "subjects": normalized_subjects,
        "invalid_subjects": invalid_subjects,
        "summary": dict(counts),
        "total_rows": len(audited_rows),
        "column_map": column_map,
        "rows": audited_rows,
    }


def parse_year(value: Optional[str]) -> Optional[int]:
    if value is None or str(value).strip() == "":
        return None
    return int(value)


def main() -> int:
    parser = argparse.ArgumentParser(description="Audit volunteer plan subject eligibility.")
    parser.add_argument("--input", required=True, type=Path, help="CSV/JSON/XLSX volunteer plan.")
    parser.add_argument("--year", help="Admission year, e.g. 2026.")
    parser.add_argument("--edition", help="Subject requirement edition: 2024 or 2027.")
    parser.add_argument("--level", required=True, help="本科 or 专科.")
    parser.add_argument("--subjects", required=True, help="Three subjects, e.g. 物理,化学,生物.")
    parser.add_argument("--out", type=Path, help="Report JSON path.")
    parser.add_argument("--db", type=Path, default=DEFAULT_DB_PATH, help="SQLite index path.")
    args = parser.parse_args()

    try:
        rows = read_input_rows(args.input)
        report = audit_rows(
            rows=rows,
            year=parse_year(args.year),
            edition=args.edition,
            level=args.level,
            subjects=args.subjects,
            db_path=args.db,
        )
    except Exception as exc:
        error_report = {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "overall_status": "ERROR",
            "hard_gate_passed": False,
            "error": str(exc),
        }
        if args.out:
            args.out.parent.mkdir(parents=True, exist_ok=True)
            with args.out.open("w", encoding="utf-8") as f:
                json.dump(error_report, f, ensure_ascii=False, indent=2)
                f.write("\n")
        print(json.dumps(error_report, ensure_ascii=False, indent=2))
        return 2

    if args.out:
        args.out.parent.mkdir(parents=True, exist_ok=True)
        with args.out.open("w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
            f.write("\n")
    print(json.dumps({"overall_status": report["overall_status"], "summary": report["summary"]}, ensure_ascii=False))
    return 0 if report["hard_gate_passed"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
